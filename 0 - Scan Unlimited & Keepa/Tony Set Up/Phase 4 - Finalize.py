import pandas as pd
import urllib.parse
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill
import math, os

# 1) Load CSV
INPUT_FILE = "Keepa Cleaned.csv"
if not os.path.exists(INPUT_FILE):
    raise FileNotFoundError(f"Could not find {INPUT_FILE}")
df = pd.read_csv(INPUT_FILE)

# 2) Define column groups
currency_cols  = ["Cost","List","Profit","30/90 BB","30/90 Profit","CPT","CPT Profit","Amz"]
percent_cols   = ["ROI","%","CPT ROI","Amz OOS"]
integer_cols   = ["Offer","Rank","# Sold","Rating","T Rating"]
hyperlink_cols = ["Source","Amazon"]
blank_zero     = ["# Sold","CPT","CPT Profit","CPT ROI","Amz"]

# 3) Convert fixed columns
for col in currency_cols:
    if col in df:
        s = df[col].astype(str)
        s = s.replace(r"\(([^)]+)\)", r"-\1", regex=True)  # “($X)” → “-X”
        s = s.replace(r"[\$,]", "", regex=True)
        df[col] = pd.to_numeric(s, errors="coerce")

for col in percent_cols:
    if col in df:
        s = df[col].astype(str)
        s = s.replace(r"\(([^)]+)\)", r"-\1", regex=True).str.rstrip("%")
        df[col] = pd.to_numeric(s, errors="coerce").div(100)

for col in integer_cols:
    if col in df:
        s = df[col].astype(str).replace(",", "", regex=True)
        df[col] = pd.to_numeric(s, errors="coerce").dropna().astype(int)

# 4) Build filename & sheet name
first_src = df.get("Source", pd.Series()).dropna().astype(str).iat[0] if "Source" in df else ""
netloc    = urllib.parse.urlparse(first_src).netloc.replace("www.","")
dom       = netloc.split(".")[0] if netloc else "Output"
if dom.lower().endswith("wss"):
    dom_title = dom[:-3].capitalize() + dom[-3:].upper()
else:
    dom_title = dom.capitalize()

date_str    = datetime.now().strftime("%m.%d.%y")
base        = f"{date_str} {dom_title}"
OUTPUT_FILE = f"{base}.xlsx"
SHEET_NAME  = base[:31]

# 5) Create workbook
wb = Workbook()
ws = wb.active
ws.title        = SHEET_NAME
ws.freeze_panes = "A2"

# 6) Write header
for idx, hdr in enumerate(df.columns, start=1):
    cell = ws.cell(row=1, column=idx, value=hdr)
    cell.font = Font(bold=True)

# 7) Write data + formatting
green = PatternFill("solid", fgColor="00FF00")
for r, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
    for c, (col, val) in enumerate(zip(df.columns, row), start=1):
        cell = ws.cell(row=r, column=c)
        # NaN → blank
        if isinstance(val, float) and math.isnan(val):
            continue
        # blank-if-zero
        if col in blank_zero and val == 0:
            continue
        # hyperlinks
        if col in hyperlink_cols:
            if val:
                cell.value     = val
                cell.hyperlink = val
                cell.style     = "Hyperlink"
            continue
        # currency
        if col in currency_cols:
            cell.value = val
            if col in ["30/90 Profit","CPT Profit"]:
                cell.number_format = "$#,##0.00;[Red]($#,##0.00)"
            else:
                cell.number_format = "$#,##0.00"
        # percent
        elif col in percent_cols:
            cell.value = val
            cell.number_format = "0%"
        # integer
        elif col in integer_cols:
            cell.value = val
            cell.number_format = "#,##0"
        # custom/other
        else:
            cell.value = val
        # conditional highlights
        if col=="Profit"   and val>10:    cell.fill = green
        if col=="ROI"      and val>0.25:  cell.fill = green
        if col=="CPT ROI"  and val>0.10:  cell.fill = green
        if col=="Amz OOS"  and val>0.70:  cell.fill = green
        if col=="BB/SUP"   and val=="BB": cell.fill = green

# 8) Column widths = max(auto_fit, minimum) with +1 padding
min_widths = {
    "ASIN":15, "Cost":7,  "List":7,  "Profit":6, "ROI":5,
    "Offer":5,"Rank":6, "# Sold":6, "30/90 BB":8, "30/90 Profit":11,
    "Rating":8, "T Rating":8, "%":5, "Return":7, "BB/SUP":7,
    "CPT":7, "CPT Profit":9, "CPT ROI":8, "Amz":6, "Amz OOS":9
}
auto_cols = [c for c in min_widths if c in df.columns]

for col in auto_cols:
    idx    = list(df.columns).index(col) + 1
    letter = get_column_letter(idx)
    max_len = len(str(col))
    for r in range(2, ws.max_row+1):
        cell = ws.cell(row=r, column=idx)
        if cell.value is None:
            continue
        # build display string
        if col in currency_cols:
            num = float(cell.value)
            if col in ["30/90 Profit","CPT Profit"] and num<0:
                disp = f"(${abs(num):,.2f})"
            else:
                disp = f"${num:,.2f}"
        elif col in percent_cols:
            num  = float(cell.value)
            disp = f"{num:.0%}"
        elif col in integer_cols:
            num  = int(cell.value)
            disp = f"{num:,}"
        else:
            disp = str(cell.value)
        max_len = max(max_len, len(disp))
    # apply max of auto and minimum, with +1 padding
    ws.column_dimensions[letter].width = max(max_len + 1, min_widths[col])

# Title fixed at 40
if "Title" in df.columns:
    i = list(df.columns).index("Title") + 1
    ws.column_dimensions[get_column_letter(i)].width = 40

# 9) Save
wb.save(OUTPUT_FILE)
print(f"✓ Saved → {OUTPUT_FILE}")
